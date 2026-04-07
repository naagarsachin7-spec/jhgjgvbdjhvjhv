import io
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from processor import clear_template_workbook, fill_template, parse_invoice_upload, records_to_df, workbook_to_bytes

APP_DIR = Path(__file__).parent
DEFAULT_TEMPLATE_PATH = APP_DIR / "template_clean.xlsx"

st.set_page_config(page_title="GST Master Data Builder", page_icon="📄", layout="wide")
st.title("GST Master Data Builder")
st.caption("Upload invoice Excel files and generate the same GST master template format automatically.")

with st.sidebar:
    st.subheader("How it works")
    st.markdown("""
1. Upload invoice Excel files  
2. Optionally upload your own master template  
3. Generate output workbook  
4. Download the filled file
""")
    st.info("This version is tuned for the current Varkesh invoice Excel format and fills the same template tabs without changing headers or tab names.")

invoice_files = st.file_uploader(
    "Upload invoice Excel files",
    type=["xlsx", "xlsm", "xltx", "xltm"],
    accept_multiple_files=True,
)

template_file = st.file_uploader(
    "Optional: upload your master template workbook",
    type=["xlsx", "xlsm", "xltx", "xltm"],
    accept_multiple_files=False,
    help="If skipped, the built-in clean template will be used.",
)

clear_existing = st.checkbox("Clear old data from template before filling", value=True)

if invoice_files:
    parsed_records = []
    errors = []
    for uploaded in invoice_files:
        try:
            parsed_records.append(parse_invoice_upload(uploaded))
        except Exception as e:
            errors.append({"file": uploaded.name, "error": str(e)})

    if parsed_records:
        st.subheader("Parsed invoice preview")
        st.dataframe(records_to_df(parsed_records), use_container_width=True)
        c1, c2, c3 = st.columns(3)
        c1.metric("Invoices parsed", len(parsed_records))
        c2.metric("Total taxable value", f"{sum(r['taxable_value'] for r in parsed_records):,.2f}")
        c3.metric("Total invoice value", f"{sum(r['total_invoice_value'] for r in parsed_records):,.2f}")

    if errors:
        st.subheader("Files with errors")
        st.dataframe(pd.DataFrame(errors), use_container_width=True)

    if st.button("Generate GST Master Workbook", type="primary", use_container_width=True, disabled=not parsed_records):
        if template_file:
            wb = load_workbook(io.BytesIO(template_file.getvalue()))
        else:
            wb = load_workbook(DEFAULT_TEMPLATE_PATH)

        if clear_existing:
            wb = clear_template_workbook(wb)

        wb = fill_template(wb, parsed_records)
        out_bytes = workbook_to_bytes(wb)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.success("Workbook generated successfully.")
        st.download_button(
            label="Download filled GST master workbook",
            data=out_bytes,
            file_name=f"GST_Master_Output_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
else:
    st.info("Upload one or more invoice Excel files to begin.")
