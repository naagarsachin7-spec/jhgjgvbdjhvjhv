import io
import re
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

STATE_CODE_TO_NAME = {
    "01": "Jammu and Kashmir", "02": "Himachal Pradesh", "03": "Punjab", "04": "Chandigarh",
    "05": "Uttarakhand", "06": "Haryana", "07": "Delhi", "08": "Rajasthan", "09": "Uttar Pradesh",
    "10": "Bihar", "11": "Sikkim", "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur",
    "15": "Mizoram", "16": "Tripura", "17": "Meghalaya", "18": "Assam", "19": "West Bengal",
    "20": "Jharkhand", "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh",
    "24": "Gujarat", "25": "Daman and Diu", "26": "Dadra and Nagar Haveli and Daman and Diu",
    "27": "Maharashtra", "29": "Karnataka", "30": "Goa", "31": "Lakshadweep", "32": "Kerala",
    "33": "Tamil Nadu", "34": "Puducherry", "35": "Andaman and Nicobar Islands", "36": "Telangana",
    "37": "Andhra Pradesh", "38": "Ladakh", "97": "Other Territory"
}
STATE_NAME_ALIASES = {
    "uttarakhand": "Uttarakhand",
    "uttrakhand": "Uttarakhand",
    "orissa": "Odisha",
    "odisha": "Odisha",
    "chattisgarh": "Chhattisgarh",
    "chhattisgarh": "Chhattisgarh",
    "madhya pradesh": "Madhya Pradesh",
    "maharashtra": "Maharashtra",
    "delhi": "Delhi",
    "haryana": "Haryana",
    "punjab": "Punjab",
    "himachal pradesh": "Himachal Pradesh",
    "jammu and kashmir": "Jammu and Kashmir",
    "j&k": "Jammu and Kashmir",
    "rajasthan": "Rajasthan",
    "uttar pradesh": "Uttar Pradesh",
    "bihar": "Bihar",
    "gujarat": "Gujarat",
    "karnataka": "Karnataka",
    "goa": "Goa",
    "kerala": "Kerala",
    "tamil nadu": "Tamil Nadu",
    "telangana": "Telangana",
    "andhra pradesh": "Andhra Pradesh",
    "west bengal": "West Bengal",
    "jharkhand": "Jharkhand",
    "assam": "Assam",
    "meghalaya": "Meghalaya",
    "tripura": "Tripura",
    "arunachal pradesh": "Arunachal Pradesh",
    "mizoram": "Mizoram",
    "manipur": "Manipur",
    "nagaland": "Nagaland",
    "sikkim": "Sikkim",
    "ladakh": "Ladakh",
    "chandigarh": "Chandigarh",
    "puducherry": "Puducherry",
}
STATE_NAME_TO_CODE = {v: k for k, v in STATE_CODE_TO_NAME.items()}

GST_MASTER_HEADERS = [
    "row_number","source_file_name","source_sheet_name","supplier_name","supplier_gstin","recipient_name",
    "recipient_gstin","recipient_registration_type","invoice_number","invoice_date","invoice_type",
    "document_type","place_of_supply","state_code","state_name","product_description","item_code",
    "hsn_sac","uqc","quantity","unit_cost","taxable_value","discount","delivery_charge","gst_rate",
    "cgst_amount","sgst_amount","igst_amount","cess_rate","cess_amount","total_invoice_value"
]


def normalize_text(value: Optional[str]) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def safe_float(value, default=0.0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except Exception:
        text = re.sub(r"[^0-9.\-]", "", str(value))
        return float(text) if text else default


def extract_gstin(text: str) -> Optional[str]:
    if not text:
        return None
    m = re.search(r"\b\d{2}[A-Z]{5}\d{4}[A-Z][A-Z\d]Z[A-Z\d]\b", text.upper())
    return m.group(0) if m else None


def find_state_from_text(text: str) -> Tuple[Optional[str], Optional[str]]:
    content = normalize_text(text).lower()
    for alias, state_name in sorted(STATE_NAME_ALIASES.items(), key=lambda x: len(x[0]), reverse=True):
        if alias in content:
            return STATE_NAME_TO_CODE.get(state_name), state_name
    return None, None


def invoice_sort_key(invoice_no: str):
    parts = re.findall(r"\d+", invoice_no or "")
    return [int(p) for p in parts] if parts else [invoice_no or ""]


def clear_sheet_range(ws, start_row: int, start_col: int = 1, end_col: Optional[int] = None):
    end_col = end_col or ws.max_column
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.value = None


def parse_invoice_file(path: str) -> Dict:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return _parse_invoice_ws(ws, source_name=path.split("/")[-1], sheet_name=wb.sheetnames[0])


def parse_invoice_upload(uploaded_file) -> Dict:
    wb = load_workbook(io.BytesIO(uploaded_file.getvalue()), data_only=True)
    ws = wb[wb.sheetnames[0]]
    return _parse_invoice_ws(ws, source_name=uploaded_file.name, sheet_name=wb.sheetnames[0])


def _parse_invoice_ws(ws, source_name: str, sheet_name: str) -> Dict:
    supplier_name = normalize_text(ws["A2"].value)
    supplier_gstin = extract_gstin(normalize_text(ws["A6"].value)) or ""

    ship_to_row = None
    invoice_no = ""
    invoice_date = None
    for row in range(1, ws.max_row + 1):
        row_values = [normalize_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        joined = " | ".join(v for v in row_values if v)
        if "Ship To" in joined:
            ship_to_row = row
        for col in range(1, ws.max_column + 1):
            val = normalize_text(ws.cell(row, col).value)
            if "Invoice No" in val:
                parts = val.split(":")
                invoice_no = normalize_text(parts[-1]) if parts else val
            if val.lower().startswith("date"):
                next_val = ws.cell(row, col + 1).value if col < ws.max_column else None
                if isinstance(next_val, datetime):
                    invoice_date = next_val

    recipient_name = normalize_text(ws.cell(ship_to_row + 1, 1).value) if ship_to_row else ""
    address_lines = []
    if ship_to_row:
        # collect address rows up to the product table header; invoice/date labels may appear in parallel columns
        for row in range(ship_to_row + 2, min(14, ws.max_row + 1)):
            first_cell = normalize_text(ws.cell(row, 1).value)
            if first_cell:
                address_lines.append(first_cell)

    address_text = ", ".join(address_lines)
    state_code, state_name = find_state_from_text(address_text)
    place_of_supply = f"{state_code}-{state_name}" if state_code and state_name else ""
    recipient_gstin = ""

    product_description = ""
    hsn_sac = ""
    quantity = 0.0
    unit_cost = 0.0
    taxable_value = 0.0
    total_invoice_value = 0.0
    delivery_charge = 0.0
    igst_amount = 0.0
    cgst_amount = 0.0
    sgst_amount = 0.0
    gst_rate = 0.0

    for row in range(1, ws.max_row + 1):
        label_b = normalize_text(ws.cell(row, 2).value)
        label_d = normalize_text(ws.cell(row, 4).value).lower()

        if row == 17:
            product_description = normalize_text(ws.cell(row, 2).value)
            unit_cost = safe_float(ws.cell(row, 3).value)
            quantity = safe_float(ws.cell(row, 4).value, default=1.0)
        if "hsn code" in label_b.lower():
            hsn_match = re.search(r"(\d{4,8})", label_b)
            hsn_sac = hsn_match.group(1) if hsn_match else ""
        if label_d == "subtotal":
            taxable_value = safe_float(ws.cell(row, 5).value)
        if "delivery charge" in label_d:
            delivery_charge = safe_float(ws.cell(row, 5).value)
        if "igst" in label_d:
            igst_amount = safe_float(ws.cell(row, 5).value)
            rate_match = re.search(r"(\d+(\.\d+)?)", normalize_text(ws.cell(row, 4).value))
            if rate_match:
                gst_rate = safe_float(rate_match.group(1))
        if label_d.replace(" ", "").startswith("cgst"):
            cgst_amount = safe_float(ws.cell(row, 5).value)
        if label_d.replace(" ", "").startswith("sgst"):
            sgst_amount = safe_float(ws.cell(row, 5).value)
        if label_d == "total":
            total_invoice_value = safe_float(ws.cell(row, 5).value)

    if not gst_rate and taxable_value:
        gst_rate = round(((igst_amount + cgst_amount + sgst_amount) / taxable_value) * 100, 2)

    tax_type = "IGST" if igst_amount > 0 else ("CGST/SGST" if (cgst_amount > 0 or sgst_amount > 0) else "")

    return {
        "source_file_name": source_name,
        "source_sheet_name": sheet_name,
        "supplier_name": supplier_name,
        "supplier_gstin": supplier_gstin,
        "recipient_name": recipient_name,
        "recipient_gstin": recipient_gstin,
        "recipient_registration_type": "Unregistered / B2C",
        "invoice_number": invoice_no,
        "invoice_date": invoice_date,
        "invoice_type": "B2C",
        "document_type": "Tax Invoice",
        "place_of_supply": place_of_supply,
        "state_code": state_code,
        "state_name": state_name,
        "product_description": product_description,
        "item_code": "",
        "hsn_sac": hsn_sac,
        "uqc": "NOS",
        "quantity": quantity,
        "unit_cost": round(unit_cost, 2),
        "taxable_value": round(taxable_value, 2),
        "discount": 0.0,
        "delivery_charge": round(delivery_charge, 2),
        "gst_rate": round(gst_rate, 2),
        "cgst_amount": round(cgst_amount, 2),
        "sgst_amount": round(sgst_amount, 2),
        "igst_amount": round(igst_amount, 2),
        "cess_rate": 0.0,
        "cess_amount": 0.0,
        "total_invoice_value": round(total_invoice_value, 2),
        "tax_type": tax_type,
        "address_text": address_text,
    }


def clear_template_workbook(wb: Workbook) -> Workbook:
    if "GST_Master_Data" in wb.sheetnames:
        clear_sheet_range(wb["GST_Master_Data"], 2, 1, wb["GST_Master_Data"].max_column)

    if "b2b,sez,de" in wb.sheetnames:
        clear_sheet_range(wb["b2b,sez,de"], 5, 1, wb["b2b,sez,de"].max_column)
        for ref in ["A3", "C3", "E3", "L3"]:
            wb["b2b,sez,de"][ref] = 0

    if "b2cs" in wb.sheetnames:
        clear_sheet_range(wb["b2cs"], 5, 1, wb["b2cs"].max_column)
        wb["b2cs"]["E3"] = 0
        wb["b2cs"]["F3"] = 0

    if "cdnr" in wb.sheetnames:
        clear_sheet_range(wb["cdnr"], 5, 1, wb["cdnr"].max_column)
        for ref in ["A3", "C3", "E3", "K3"]:
            wb["cdnr"][ref] = 0

    if "hsn(b2b)" in wb.sheetnames:
        clear_sheet_range(wb["hsn(b2b)"], 5, 1, wb["hsn(b2b)"].max_column)
        for ref in ["A3", "E3", "G3", "H3", "I3", "J3", "K3"]:
            wb["hsn(b2b)"][ref] = 0

    if "hsn(b2c)" in wb.sheetnames:
        clear_sheet_range(wb["hsn(b2c)"], 5, 1, wb["hsn(b2c)"].max_column)
        for ref in ["A3", "E3", "G3", "H3", "I3", "J3", "K3"]:
            wb["hsn(b2c)"][ref] = 0

    if "docs" in wb.sheetnames:
        clear_sheet_range(wb["docs"], 5, 1, wb["docs"].max_column)
        wb["docs"]["D3"] = 0
        wb["docs"]["E3"] = 0

    return wb


def fill_template(wb: Workbook, records: List[Dict]) -> Workbook:
    wb = clear_template_workbook(wb)

    ws = wb["GST_Master_Data"]
    header_map = {normalize_text(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    for idx, record in enumerate(records, start=1):
        row = idx + 1
        record_with_row = {"row_number": idx, **record}
        for key in GST_MASTER_HEADERS:
            if key in header_map:
                ws.cell(row, header_map[key]).value = record_with_row.get(key)

    b2cs_groups = {}
    total_cess = 0.0
    for record in records:
        if record["recipient_gstin"]:
            continue
        key = ("OE", record["place_of_supply"], None, record["gst_rate"], None)
        b2cs_groups.setdefault(key, {"taxable_value": 0.0, "cess_amount": 0.0})
        b2cs_groups[key]["taxable_value"] += record["taxable_value"]
        b2cs_groups[key]["cess_amount"] += record["cess_amount"]
        total_cess += record["cess_amount"]

    ws = wb["b2cs"]
    current_row = 5
    total_taxable_value = 0.0
    for (typ, pos, app_tax, rate, ecommerce), vals in sorted(b2cs_groups.items(), key=lambda x: x[0][1] or ""):
        ws.cell(current_row, 1).value = typ
        ws.cell(current_row, 2).value = pos
        ws.cell(current_row, 3).value = app_tax
        ws.cell(current_row, 4).value = rate
        ws.cell(current_row, 5).value = round(vals["taxable_value"], 2)
        ws.cell(current_row, 6).value = round(vals["cess_amount"], 2)
        ws.cell(current_row, 7).value = ecommerce
        total_taxable_value += vals["taxable_value"]
        current_row += 1
    ws["E3"] = round(total_taxable_value, 2)
    ws["F3"] = round(total_cess, 2)

    hsn_groups = {}
    for record in records:
        if record["recipient_gstin"]:
            continue
        key = (record["hsn_sac"], record["product_description"], record["uqc"], record["gst_rate"])
        hsn_groups.setdefault(key, {
            "qty": 0.0, "total_value": 0.0, "taxable": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0
        })
        hsn_groups[key]["qty"] += record["quantity"]
        hsn_groups[key]["total_value"] += record["total_invoice_value"]
        hsn_groups[key]["taxable"] += record["taxable_value"]
        hsn_groups[key]["igst"] += record["igst_amount"]
        hsn_groups[key]["cgst"] += record["cgst_amount"]
        hsn_groups[key]["sgst"] += record["sgst_amount"]
        hsn_groups[key]["cess"] += record["cess_amount"]

    ws = wb["hsn(b2c)"]
    current_row = 5
    totals = {"total_value": 0.0, "taxable": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
    for (hsn, desc, uqc, rate), vals in sorted(hsn_groups.items(), key=lambda x: (x[0][0] or "", x[0][1] or "")):
        ws.cell(current_row, 1).value = hsn
        ws.cell(current_row, 2).value = desc
        ws.cell(current_row, 3).value = uqc
        ws.cell(current_row, 4).value = round(vals["qty"], 2)
        ws.cell(current_row, 5).value = round(vals["total_value"], 2)
        ws.cell(current_row, 6).value = rate
        ws.cell(current_row, 7).value = round(vals["taxable"], 2)
        ws.cell(current_row, 8).value = round(vals["igst"], 2)
        ws.cell(current_row, 9).value = round(vals["cgst"], 2)
        ws.cell(current_row, 10).value = round(vals["sgst"], 2)
        ws.cell(current_row, 11).value = round(vals["cess"], 2)
        totals["total_value"] += vals["total_value"]
        totals["taxable"] += vals["taxable"]
        totals["igst"] += vals["igst"]
        totals["cgst"] += vals["cgst"]
        totals["sgst"] += vals["sgst"]
        totals["cess"] += vals["cess"]
        current_row += 1
    ws["A3"] = len(hsn_groups)
    ws["E3"] = round(totals["total_value"], 2)
    ws["G3"] = round(totals["taxable"], 2)
    ws["H3"] = round(totals["igst"], 2)
    ws["I3"] = round(totals["cgst"], 2)
    ws["J3"] = round(totals["sgst"], 2)
    ws["K3"] = round(totals["cess"], 2)

    ws = wb["docs"]
    sorted_invoices = sorted([r["invoice_number"] for r in records if r["invoice_number"]], key=invoice_sort_key)
    ws["D3"] = len(sorted_invoices)
    ws["E3"] = 0
    if sorted_invoices:
        ws["A5"] = "Invoices for outward supply"
        ws["B5"] = sorted_invoices[0]
        ws["C5"] = sorted_invoices[-1]
        ws["D5"] = len(sorted_invoices)
        ws["E5"] = 0

    return wb


def records_to_df(records: List[Dict]) -> pd.DataFrame:
    rows = []
    for idx, record in enumerate(records, start=1):
        row = {"row_number": idx}
        row.update(record)
        rows.append(row)
    df = pd.DataFrame(rows)
    if not df.empty:
        cols = [
            "row_number","source_file_name","recipient_name","invoice_number","invoice_date",
            "place_of_supply","product_description","hsn_sac","taxable_value","igst_amount",
            "delivery_charge","total_invoice_value"
        ]
        df = df[[c for c in cols if c in df.columns]]
    return df


def workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
